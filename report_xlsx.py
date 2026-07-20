"""Аннотированный экспорт результатов прогона в xlsx.

Подсвечивает проблемные ячейки цветом и вешает подробные комментарии, плюс
лист «Легенда и проблемы» с расшифровкой. Используется кнопкой скачивания в
app.py и headless-сборкой. Строится на openpyxl (полноценные комментарии).

Типы проблем:
  • закуп: поставщик не классифицирован (строка возвращена в свод по supp_raw);
  • гейт: Δ ≠ 0 (рассинхрон закуп ↔ Загружено);
  • движение: отрицательный Остаток_конец (продано больше стока — нет carry?);
  • ручная сверка: раскраска по категории действия.
"""
from __future__ import annotations

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# палитра
C_HDR = PatternFill("solid", fgColor="1F3864")
C_PROBLEM = PatternFill("solid", fgColor="FFC9A3")   # оранжевый — неклассиф. поставщик
C_GATE_BAD = PatternFill("solid", fgColor="FF9E9E")  # красный — расхождение гейта
C_NEG = PatternFill("solid", fgColor="FF9E9E")       # красный — отриц. остаток
C_LISTING = PatternFill("solid", fgColor="FFF3B0")   # жёлтый — сопоставить листинг
C_REGION = PatternFill("solid", fgColor="FFC9A3")    # оранжевый — задать регион
C_CARRY = PatternFill("solid", fgColor="D9D9D9")     # серый — carry-шум
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)

CAT_FILL = {"Сопоставить листинг": C_LISTING,
            "Задать регион (склад)": C_REGION,
            "Вероятно снимется carry-over": C_CARRY}
CAT_NOTE = {
    "Сопоставить листинг": "Листинг маркетплейса не найден в mappings/. "
        "Подтвердите catID в окне приложения и обновите маппинг — тогда позиция "
        "будет разноситься автоматически.",
    "Задать регион (склад)": "Регион ключа не выводится из выгрузок (складской "
        "факт). Требуется ручная привязка региона/раздел склада — это потолок "
        "автоматического сплита.",
    "Вероятно снимется carry-over": "Скорее всего снимется, если подать "
        "Остаток_нач (перенос прошлого месяца): не хватило входящего стока для "
        "разнесения спроса.",
}


def _write(ws, df: pd.DataFrame, widths=None):
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(1, c)
        cell.fill = C_HDR; cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center")
    for _, row in df.iterrows():
        vals = []
        for v in row.values:
            try:
                if pd.isna(v):
                    vals.append(None); continue
            except (TypeError, ValueError):
                pass
            vals.append(v.item() if hasattr(v, "item") else v)
        ws.append(vals)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(df.columns) + 1):
            ws.cell(r, c).font = BODY_FONT
    for i in range(len(df.columns)):
        ws.column_dimensions[get_column_letter(i + 1)].width = (widths or {}).get(
            df.columns[i], 18)
    ws.freeze_panes = "A2"


def _col(df, name):
    return list(df.columns).index(name) + 1


def _legend(wb, problems: list):
    ws = wb.create_sheet("Легенда и проблемы", 0)
    ws["A1"] = "Легенда цветов и найденные проблемы"
    ws["A1"].font = Font(bold=True, size=12, name="Arial")
    ws.append([])
    legend = [
        ("Оранжевый", "FFC9A3", "Поставщик не классифицирован / регион не задан"),
        ("Красный", "FF9E9E", "Расхождение гейта или отрицательный остаток"),
        ("Жёлтый", "FFF3B0", "Листинг не сопоставлен с каталогом"),
        ("Серый", "D9D9D9", "Вероятно снимется при подаче Остаток_нач"),
    ]
    ws.append(["Цвет", "", "Значение"])
    for r in (3,):
        for c in (1, 3):
            ws.cell(r, c).font = Font(bold=True, name="Arial")
    for name, hexc, desc in legend:
        ws.append([name, "", desc])
        cell = ws.cell(ws.max_row, 2)
        cell.fill = PatternFill("solid", fgColor=hexc)
    ws.append([])
    ws.append(["Найденные проблемы:"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, name="Arial")
    for p in problems:
        ws.append([p])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 90


def build(res: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    problems = []

    # ── ГЕЙТ ──
    if "reconcile" in res:
        rec = res["reconcile"]
        ws = wb.create_sheet("Гейт")
        _write(ws, rec)
        dc = _col(rec, "Δ")
        bad = 0
        for i, (_, row) in enumerate(rec.iterrows(), start=2):
            if row["Юнит"] != "—" and row["Δ"] != 0:
                bad += 1
                cell = ws.cell(i, dc)
                cell.fill = C_GATE_BAD
                cell.comment = Comment(
                    f"Расхождение Δ={row['Δ']}: Загружено движения ≠ закуп зоны "
                    f"«{row['Зона']}». Проверьте проводку/фильтры.", "SalesFlow")
        problems.append("Гейт: " + ("расхождений нет (Δ=0 по всем юнитам)."
                        if bad == 0 else f"расхождение в {bad} юнит(ах) — см. лист «Гейт»."))

    # ── ЗАКУП: свод + по площадкам, неклассиф. строки возвращены и подсвечены ──
    if "purchases" in res:
        pur = res["purchases"]
        unc_total = sum(int(v["unclassified"]["Количество"].sum())
                        for v in pur.values() if len(v["unclassified"]))
        # сводный
        parts = []
        for z, v in pur.items():
            f = v["flat"].copy()
            u = v["unclassified"].copy()
            parts += [f, u]
        svod = pd.concat([p for p in parts if len(p)], ignore_index=True) if parts else pd.DataFrame()
        _purchase_sheet(wb, "Закуп (свод)", svod, pur, combined=True)
        for z, v in pur.items():
            df = pd.concat([d for d in (v["flat"], v["unclassified"]) if len(d)],
                           ignore_index=True)
            if len(df):
                _purchase_sheet(wb, f"Закуп {z}"[:31], df, {z: v})
        if unc_total:
            problems.append(
                f"Закуп: {unc_total} ключей от неклассифицированных поставщиков "
                f"возвращены в свод (подсвечены оранжевым). Данные не теряются, "
                f"но поставщики требуют добавления в classify_supplier.")
        else:
            problems.append("Закуп: все поставщики классифицированы.")

    # ── ПРОДАЖИ + ДВИЖЕНИЕ ──
    sales = res.get("sales", {})
    for ch, out in sales.items():
        if "sales_multi" in out:
            for name, s in out["sales_multi"].items():
                _write(wb.create_sheet(f"Продажи {name}"[:31]), s)
        elif "sales" in out:
            _write(wb.create_sheet(f"Продажи {ch}"[:31]), out["sales"])
        mv = out["movement"]
        ws = wb.create_sheet(f"Движение {ch}"[:31])
        _write(ws, mv)
        if "Остаток_конец" in mv.columns:
            kc = _col(mv, "Остаток_конец")
            neg = 0
            for i, (_, row) in enumerate(mv.iterrows(), start=2):
                val = pd.to_numeric(pd.Series([row["Остаток_конец"]]), errors="coerce").iloc[0]
                if pd.notna(val) and val < 0:
                    neg += 1
                    cell = ws.cell(i, kc); cell.fill = C_NEG
                    cell.comment = Comment(
                        "Отрицательный остаток: продано больше, чем было в стоке. "
                        "Обычно значит нехватку Остаток_нач (перенос) или Событий.",
                        "SalesFlow")
            if neg:
                problems.append(f"Движение {ch}: {neg} позиц. с отрицательным "
                                f"Остаток_конец (не хватает Остаток_нач/Событий).")

    # ── РУЧНАЯ СВЕРКА (раскраска по категории) ──
    if "review" in res and len(res["review"]):
        rv = res["review"]
        ws = wb.create_sheet("Ручная сверка")
        _write(ws, rv)
        cc = _col(rv, "Категория")
        for i, (_, row) in enumerate(rv.iterrows(), start=2):
            cat = row["Категория"]
            fill = CAT_FILL.get(cat)
            if fill:
                cell = ws.cell(i, cc); cell.fill = fill
                cell.comment = Comment(CAT_NOTE.get(cat, ""), "SalesFlow")
        by = rv.groupby("Категория")["Нераспределено"].agg(["size", "sum"])
        problems.append("Ручная сверка: " + "; ".join(
            f"{c} — {int(r['size'])} строк / {int(r['sum'])} нераспр." for c, r in by.iterrows()))

    # ── ПЕРЕНОС ОСТАТКОВ ──
    if sales:
        carry = []
        for ch, out in sales.items():
            m = out["movement"][["ID", "Название", "Регион", "Остаток_конец"]].copy()
            m = m.rename(columns={"Остаток_конец": "Остаток_нач"})
            m.insert(0, "Канал", ch)
            carry.append(m)
        _write(wb.create_sheet("Перенос остатков"),
               pd.concat(carry, ignore_index=True))

    _legend(wb, problems or ["Проблем не обнаружено."])
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def _purchase_sheet(wb, title, df, pur, combined=False):
    """Лист закупа: неклассиф. строки (Поставщик отсутствует в classify_supplier)
    подсвечиваются оранжевым, на ячейку Поставщик вешается комментарий."""
    # множество (площадка, pid, supp) неклассифицированных — для подсветки
    unc = set()
    for z, v in pur.items():
        for _, r in v["unclassified"].iterrows():
            unc.add((r["Площадка"], int(r["ID продукта"]), str(r["Поставщик"])))
    ws = wb.create_sheet(title)
    _write(ws, df, widths={"Название": 40, "Поставщик": 26})
    ps = _col(df, "Поставщик")
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        key = (row["Площадка"], int(row["ID продукта"]), str(row["Поставщик"]))
        if key in unc:
            cell = ws.cell(i, ps); cell.fill = C_PROBLEM
            cell.comment = Comment(
                f"Поставщик «{row['Поставщик']}» не распознан classify_supplier. "
                f"Строка возвращена в свод по сырому имени, чтобы не терять "
                f"{int(row['Количество'])} ключей. Добавьте поставщика в "
                f"classify_supplier (KeyFlow) для корректной группировки.", "SalesFlow")

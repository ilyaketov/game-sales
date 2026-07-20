"""
Движок автоматизации отчётов закупа: преобразует сырые отчёты биллинга
в формат ЗАКУП (свод) для загрузки в QuickBooks.

Использование:
    from engine import Pipeline
    p = Pipeline(report1_path, report2_path, genba_path)
    aggregated = p.aggregate("Plati")
    p.save_to_excel(aggregated, "Plati", "out/Plati_zakup.xlsx")
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config import (
    COLS_GENBA,
    COLS_R1,
    COLS_R2,
    CNY_SUPPLIERS,
    DEFAULT_B2B_ZONE,
    DEFAULT_CURRENCY,
    DUAL_USE_PARTNERS,
    FX_FALLBACK_DEFAULT,
    GB_SUPPLIER_MARKERS,
    CHINAPLAY_SUPPLIER_MARKERS,
    PARTNER_TO_ZONE,
    PLOSHADKA_MAP,
    RUB_CNY_RATE,
    SUPPLIER_CURRENCY,
    SUPPLIER_EXACT_RULES,
    SUPPLIER_MAPPING,
    SUPPLIER_SUBSTRING_RULES,
    SYNTH_PLOSHADKA,
)

# Префикс типа "RUB GAMES." / "USD (CIS) GAMES." / "EUR GAMES (Tier 3)."
_PREFIX_RE = re.compile(r"^\w{3}\s+(?:\(\w+\)\s+)?GAMES(?:\s*\([^)]*\))?\.\s*")
# Хвост типа "Green Man Gaming 2" → удаляем " 2"
_TRAILING_NUM_RE = re.compile(r"\s+\d+$")


# ===========================================================================
# Классификатор поставщиков
# ===========================================================================
def classify_supplier(raw_name: str | None) -> str | None:
    """Превращает сырое имя поставщика из биллинга в имя группы для ЗАКУП (свод).

    Возвращает None для пустого/неизвестного входа — это сигнал, что нужно
    добавить запись в SUPPLIER_MAPPING (config.py).
    """
    if raw_name is None or pd.isna(raw_name):
        return None
    name = str(raw_name).strip()
    if not name:
        return None

    # 1. Спецсопоставления по подстрокам (Genba/Epay/PID)
    for substring, group in SUPPLIER_SUBSTRING_RULES:
        if substring in name:
            return group

    # 2. Точные совпадения (без префикса)
    if name in SUPPLIER_EXACT_RULES:
        return SUPPLIER_EXACT_RULES[name]

    # 3. Снимаем префикс и суффикс-число, ищем в словаре
    core = _PREFIX_RE.sub("", name).strip()
    core_no_num = _TRAILING_NUM_RE.sub("", core).strip()

    return (
        SUPPLIER_MAPPING.get(core_no_num)
        or SUPPLIER_MAPPING.get(core)
        or SUPPLIER_MAPPING.get(name)
    )


# ===========================================================================
# Вычисление площадки из (Партнёр, Ключ куплен в сток, Поставщик)
# ===========================================================================
def compute_ploshadka_row(partner, in_stock, supplier, prod_ccy=None) -> str | None:
    """Возвращает имя площадки для одной строки R1/R2 — или None, если строка
    не относится к закупочной зоне (это просто обычная продажа, она в свод не входит).

    Главное правило: 'Ключ куплен в сток' решает тип операции.

    - in_stock = НЕТ: закупка у маркетплейса/стока для нашего стока.
        Партнёр обязательно технический (MP_*, StockB2B, StockGB, StockChinaPlay,
        ChinaSteamPY, ChinaPlayTaoBao, CostChinaplay, CostGB). Возвращаем zone_zakup.
        Если партнёр — обычный B2B-клиент ((CIS) Mediagame, Rokky Platform, …),
        строка с in_stock=НЕТ — это просто внешняя продажа, в закуп не идёт → None.

    - in_stock = ДА: ключ был у нас в стоке и отгружен.
        * Партнёры MP_* / технические из PARTNER_TO_ZONE → zone_peremeshchenie.
        * DUAL_USE_PARTNERS (Rokky Platform, (CN) yueshangshuma) — выбор по поставщику/валюте.
        * Прочие (обычные B2B-клиенты) → Продажи б2б.
    """
    if partner is None or (isinstance(partner, float) and pd.isna(partner)):
        return None
    p = str(partner).strip()
    if not p:
        return None

    is_stock = (
        in_stock is not None
        and not (isinstance(in_stock, float) and pd.isna(in_stock))
        and str(in_stock).strip().upper() == "ДА"
    )

    mapping = PARTNER_TO_ZONE.get(p)
    if mapping is not None:
        zakup, peremeshchenie = mapping
        return peremeshchenie if is_stock else zakup

    # Партнёр не из технического списка → закуп возможен только если ключ из нашего стока.
    if not is_stock:
        return None

    # in_stock = ДА у нетехнического партнёра: продаём из стока.
    if p in DUAL_USE_PARTNERS:
        sup = str(supplier) if supplier is not None and not (isinstance(supplier, float) and pd.isna(supplier)) else ""
        ccy = str(prod_ccy) if prod_ccy is not None and not (isinstance(prod_ccy, float) and pd.isna(prod_ccy)) else ""
        # ChinaPlay имеет приоритет над GB: 'Enaza-Games (Stock)' в CNY уходит сюда,
        # а не в продажи гб, потому что суффикс '(Stock)' + CNY указывают на ChinaPlay-сток.
        if ccy == "CNY":
            for marker in CHINAPLAY_SUPPLIER_MARKERS:
                if marker in sup:
                    return "продажи на чайнаплей"
        for marker in GB_SUPPLIER_MARKERS:
            if marker in sup:
                return "продажи гб"
        return DEFAULT_B2B_ZONE

    # Все прочие B2B-клиенты, продающие из стока
    return DEFAULT_B2B_ZONE


def add_ploshadka_column(df: pd.DataFrame, cols: dict) -> pd.DataFrame:
    """Добавляет к df колонку SYNTH_PLOSHADKA на основе Партнёр + Ключ куплен в сток + Поставщик.

    По решению v13 — площадка ВСЕГДА вычисляется, даже если в файле есть ручная колонка
    'площадка'. Это обеспечивает воспроизводимость (один и тот же файл — один и тот же
    результат) и устраняет зависимость от качества ручной разметки. Если ручной столбец
    в файле — он попадает в DataFrame как обычная колонка с исходным именем, но в зоны
    свода идёт всегда _ploshadka.
    """
    if df.empty:
        return df

    partner_col  = cols.get("partner")
    in_stock_col = cols.get("in_stock")
    supplier_col = cols.get("supplier")
    prod_ccy_col = cols.get("prod_ccy")

    if not (partner_col and partner_col in df.columns):
        # Защита от файла без колонки Партнёр — без неё классификация невозможна.
        df = df.copy()
        df[SYNTH_PLOSHADKA] = None
        return df

    partners  = df[partner_col]
    in_stock  = df[in_stock_col] if in_stock_col and in_stock_col in df.columns else pd.Series([None] * len(df), index=df.index)
    suppliers = df[supplier_col] if supplier_col and supplier_col in df.columns else pd.Series([None] * len(df), index=df.index)
    prod_ccy  = df[prod_ccy_col] if prod_ccy_col and prod_ccy_col in df.columns else pd.Series([None] * len(df), index=df.index)
    df = df.copy()
    df[SYNTH_PLOSHADKA] = [
        compute_ploshadka_row(pa, st, su, cc)
        for pa, st, su, cc in zip(partners, in_stock, suppliers, prod_ccy)
    ]
    return df


# ===========================================================================
# Структура результата валидации
# ===========================================================================
@dataclass
class ValidationResult:
    unmapped_suppliers: dict  # {raw_name: row_count}
    available_ploshadki: dict  # {ploshadka_key: total_rows}

    @property
    def is_ok(self) -> bool:
        return not self.unmapped_suppliers


# ===========================================================================
# Основной пайплайн
# ===========================================================================
class Pipeline:
    """Загружает сырые отчёты и собирает агрегаты по площадкам."""

    def __init__(
        self,
        report1_path: str | Path | None = None,
        report2_path: str | Path | None = None,
        genba_path: str | Path | None = None,
    ):
        self.df1 = self._load_r1(report1_path) if report1_path else pd.DataFrame()
        self.df2 = self._load_r2(report2_path) if report2_path else pd.DataFrame()
        self.genba = self._load_genba(genba_path) if genba_path else pd.DataFrame()

    @staticmethod
    def _pick_sheet_with_cols(path: str | Path, required: list[str]) -> str | int:
        """Возвращает имя первого листа, где присутствуют все required-колонки.

        Сравнение нечувствительно к регистру и хвостовым пробелам, чтобы
        выгрузки с переименованной колонкой ('площадка' → 'Площадка') работали.
        Если совпадения нет — возвращает 0 (первый лист) как фолбэк.
        """
        try:
            xls = pd.ExcelFile(path, engine="calamine")
        except (ImportError, ValueError):
            xls = pd.ExcelFile(path)
        for sh in xls.sheet_names:
            try:
                head = pd.read_excel(xls, sheet_name=sh, nrows=1)
            except Exception:
                continue
            cols_norm = {str(c).strip().lower() for c in head.columns}
            req_norm = {r.strip().lower() for r in required}
            if req_norm.issubset(cols_norm):
                return sh
        return xls.sheet_names[0]

    @staticmethod
    def _resolve_cols(actual_cols: list, mapping: dict) -> dict:
        """Возвращает обновлённый словарь mapping, где значения заменены на
        фактические имена колонок из actual_cols (с учётом регистра и пробелов).

        Match идёт по нормализованной форме: strip().lower(). Если ничего не
        нашлось — оставляем оригинальное значение (чтобы дальнейший usecols
        выдал понятную ошибку).
        """
        actual_by_norm = {str(c).strip().lower(): str(c) for c in actual_cols}
        resolved = {}
        for k, v in mapping.items():
            resolved[k] = actual_by_norm.get(str(v).strip().lower(), v)
        return resolved

    @staticmethod
    def _load_r1(path) -> pd.DataFrame:
        """Загрузка Universal Report 1.

        Колонка площадки больше не требуется: она вычисляется из
        Партнёр + Ключ куплен в сток. Старые выгрузки с ручной 'площадка ' / 'Площадка'
        тоже поддерживаются (через фолбэк в add_ploshadka_column).
        Колонка 'Ключ куплен в сток' могла отсутствовать в старых выгрузках —
        в этом случае все строки трактуются как закуп (in_stock = NaN → НЕТ).

        Лист данных выбирается по колонкам (pid+supplier+partner), как в R2/genba:
        начиная с июня 2026 перед сырым листом 'выгрузка' появляются пивоты
        ('закуп', 'перемещение', …), и жёсткое чтение нулевого листа давало
        пустой df1 → весь закуп из R1 молча терялся.
        """
        required = [COLS_R1["pid"], COLS_R1["supplier"], COLS_R1["partner"]]
        sheet = Pipeline._pick_sheet_with_cols(path, required)
        try:
            head = pd.read_excel(path, sheet_name=sheet, engine="calamine", nrows=1)
        except (ImportError, ValueError):
            head = pd.read_excel(path, sheet_name=sheet, nrows=1)
        actual_cols = list(head.columns)
        resolved = Pipeline._resolve_cols(actual_cols, COLS_R1)
        COLS_R1.update(resolved)

        # Берём только те ключи, для которых колонка реально есть в файле.
        actual_norm = {str(c).strip().lower() for c in actual_cols}
        cols = []
        for v in COLS_R1.values():
            if str(v).strip().lower() in actual_norm:
                cols.append(v)
        cols = list(set(cols))
        try:
            df = pd.read_excel(path, sheet_name=sheet, engine="calamine", usecols=cols)
        except (ImportError, ValueError):
            df = pd.read_excel(path, sheet_name=sheet, usecols=cols)
        return add_ploshadka_column(df, COLS_R1)

    @staticmethod
    def _load_r2(path) -> pd.DataFrame:
        """Загрузка Universal Report shipped. Терпима к доп. титульным листам
        и к мелким различиям регистра в именах колонок. Колонка площадки
        вычисляется (с фолбэком на ручную колонку из старого формата).
        """
        # Только обязательные колонки для выбора правильного листа.
        required = [COLS_R2["pid"], COLS_R2["supplier"], COLS_R2["partner"]]
        sheet = Pipeline._pick_sheet_with_cols(path, required)
        try:
            head = pd.read_excel(path, sheet_name=sheet, engine="calamine", nrows=1)
        except (ImportError, ValueError):
            head = pd.read_excel(path, sheet_name=sheet, nrows=1)
        actual_cols = list(head.columns)
        resolved = Pipeline._resolve_cols(actual_cols, COLS_R2)
        COLS_R2.update(resolved)

        actual_norm = {str(c).strip().lower() for c in actual_cols}
        cols = []
        for v in COLS_R2.values():
            if str(v).strip().lower() in actual_norm:
                cols.append(v)
        cols = list(set(cols))
        try:
            df = pd.read_excel(path, sheet_name=sheet, engine="calamine", usecols=cols)
        except (ImportError, ValueError):
            df = pd.read_excel(path, sheet_name=sheet, usecols=cols)
        return add_ploshadka_column(df, COLS_R2)

    @staticmethod
    def _load_genba(path) -> pd.DataFrame:
        """Загрузка genbaFile. Резолв имён колонок case-insensitive:
        начиная с апреля 2026 колонка 'площадка' переименована в 'Площадка'.
        """
        required = [COLS_GENBA["pid"], COLS_GENBA["qty"], COLS_GENBA["grand_total"]]
        sheet = Pipeline._pick_sheet_with_cols(path, required)
        try:
            head = pd.read_excel(path, sheet_name=sheet, engine="calamine", nrows=1)
        except (ImportError, ValueError):
            head = pd.read_excel(path, sheet_name=sheet, nrows=1)
        resolved = Pipeline._resolve_cols(list(head.columns), COLS_GENBA)
        COLS_GENBA.update(resolved)

        cols = list({v for v in COLS_GENBA.values()})
        try:
            return pd.read_excel(path, sheet_name=sheet, engine="calamine", usecols=cols)
        except (ImportError, ValueError):
            return pd.read_excel(path, sheet_name=sheet, usecols=cols)

    # -----------------------------------------------------------------------
    # Валидация: ищем неизвестные имена поставщиков по всем площадкам
    # -----------------------------------------------------------------------
    def validate(self) -> ValidationResult:
        all_supps_to_check = []

        if not self.df1.empty:
            df1f = self.df1[
                self.df1[SYNTH_PLOSHADKA]
                .astype(str).str.strip().str.lower()
                .str.startswith("закуп")
            ]
            all_supps_to_check.append(df1f[COLS_R1["supplier"]])
        if not self.df2.empty:
            df2f = self.df2[
                self.df2[SYNTH_PLOSHADKA]
                .astype(str).str.strip().str.lower()
                .str.startswith("закуп")
            ]
            all_supps_to_check.append(df2f[COLS_R2["supplier"]])

        unmapped = {}
        if all_supps_to_check:
            all_supps = pd.concat(all_supps_to_check)
            for raw, count in all_supps.value_counts().items():
                if classify_supplier(raw) is None:
                    unmapped[str(raw)] = int(count)

        # Какие площадки реально есть в данных
        available = {}
        for key in PLOSHADKA_MAP:
            count = self._row_count_for(key)
            if count > 0:
                available[key] = count

        return ValidationResult(
            unmapped_suppliers=unmapped,
            available_ploshadki=available,
        )

    @staticmethod
    def _normalize_filter(f) -> list:
        """Превращает строку или список фильтров в список (lowercase)."""
        if f is None:
            return []
        if isinstance(f, str):
            return [f.lower()]
        return [s.lower() for s in f]

    def _row_count_for(self, ploshadka_key: str) -> int:
        cfg = PLOSHADKA_MAP[ploshadka_key]
        n = 0
        r1_filters = self._normalize_filter(cfg["r1"])
        r2_filters = self._normalize_filter(cfg["r2"])
        if r1_filters and not self.df1.empty:
            mask = (self.df1[SYNTH_PLOSHADKA].astype(str).str.strip().str.lower()
                    .isin(r1_filters))
            n += int(mask.sum())
        if r2_filters and not self.df2.empty:
            mask = (self.df2[SYNTH_PLOSHADKA].astype(str).str.strip().str.lower()
                    .isin(r2_filters))
            n += int(mask.sum())
        return n

    # -----------------------------------------------------------------------
    # Агрегация для одной площадки
    # -----------------------------------------------------------------------
    def aggregate(self, ploshadka_key: str) -> pd.DataFrame:
        if ploshadka_key not in PLOSHADKA_MAP:
            raise ValueError(f"Неизвестная площадка: {ploshadka_key}")
        cfg = PLOSHADKA_MAP[ploshadka_key]

        rows = []
        for f in self._normalize_filter(cfg["r1"]):
            if not self.df1.empty:
                rows.extend(self._extract_r1_rows(f))
        for f in self._normalize_filter(cfg["r2"]):
            if not self.df2.empty:
                rows.extend(self._extract_r2_rows(f))

        # Доп. строки по подстрокам поставщика (для PLAION-перемещений и т.п.)
        extra_subs = cfg.get("extra_supplier_substrings") or []
        if extra_subs and not self.df2.empty:
            # уже извлечённые ключи (площадка, поставщик, заказ) — чтобы не задвоить
            seen_pids_supps = {(r["pid"], r["supp"]) for r in rows}
            for sub in extra_subs:
                rows.extend(self._extract_r2_by_supplier_substring(sub, exclude_ploshadkas=self._normalize_filter(cfg["r2"])))

        combined = pd.DataFrame(rows)
        if combined.empty:
            return pd.DataFrame()

        combined["supplier_group"] = combined["supp"].apply(classify_supplier)
        # RUB-сумма для CNY-поставщиков
        combined["rub_amount"] = combined.apply(
            lambda r: r["base_amount"] if r["base_ccy"] == "RUB" else None, axis=1
        )

        # Genba lookup для этой площадки
        genba_lookup = self._build_genba_lookup(cfg["genba"]) if cfg["genba"] else {}
        # Сводный курс по валютам — фолбэк для строк с пустым 'Курс фиксации'
        fx_lookup = self._fx_lookup()

        agg = (
            combined.groupby(["supplier_group", "pid"], dropna=False)
            .agg(
                qty=("qty", "sum"),
                sum_base=("base_amount", "sum"),
                sum_rub=("rub_amount", "sum"),
                sum_prod=("prod_amount", "sum"),
                prod_ccy=("prod_ccy", lambda s: s.dropna().mode().iloc[0] if s.dropna().any() else None),
                prod_name=("prod_name", "first"),
                supp_raw=("supp", "first"),
            )
            .reset_index()
        )

        agg[["unit_price", "currency"]] = agg.apply(
            lambda r: pd.Series(self._compute_price(r, genba_lookup, fx_lookup)), axis=1
        )
        agg["cost"] = agg["qty"] * agg["unit_price"]
        return agg

    def _extract_r1_rows(self, filter_value):
        df = self.df1[
            self.df1[SYNTH_PLOSHADKA].astype(str).str.strip().str.lower()
            == filter_value.lower()
        ]
        rows = []
        for _, row in df.iterrows():
            rows.append({
                "supp":        row[COLS_R1["supplier"]],
                "pid":         row[COLS_R1["pid"]],
                "prod_name":   row[COLS_R1["prod_name"]],
                "qty":         1,
                "base_amount": row[COLS_R1["base_amount"]],
                "base_ccy":    row.get(COLS_R1["base_ccy"]),
                "prod_amount": row.get(COLS_R1["prod_amount"]),
                "prod_ccy":    row.get(COLS_R1["prod_ccy"]),
                "fx_rate":     None,  # в R1 нет колонки курса
            })
        return rows

    def _extract_r2_rows(self, filter_value):
        df = self.df2[
            self.df2[SYNTH_PLOSHADKA].astype(str).str.strip().str.lower()
            == filter_value.lower()
        ]
        rows = []
        for _, row in df.iterrows():
            qty = row[COLS_R2["qty"]]
            # calamine может вернуть float — нормализуем в int
            try:
                qty = int(qty) if pd.notna(qty) else 0
            except (ValueError, TypeError):
                qty = 0
            rows.append({
                "supp":        row[COLS_R2["supplier"]],
                "pid":         row[COLS_R2["pid"]],
                "prod_name":   row[COLS_R2["prod_name"]],
                "qty":         qty,
                "base_amount": row[COLS_R2["base_amount"]],
                "base_ccy":    row.get(COLS_R2["base_ccy"]),
                "prod_amount": row.get(COLS_R2["prod_amount"]),
                "prod_ccy":    row.get(COLS_R2["prod_ccy"]),
                "fx_rate":     row.get(COLS_R2["fx_rate"]),
            })
        return rows

    def _extract_r2_by_supplier_substring(self, substring, exclude_ploshadkas=None):
        """Извлекает R2-строки, где имя поставщика содержит подстроку,
        исключая площадки, уже учтённые основным фильтром (чтобы не задвоить).
        """
        if self.df2.empty:
            return []
        mask = self.df2[COLS_R2["supplier"]].astype(str).str.contains(substring, na=False)
        if exclude_ploshadkas:
            excl = [s.lower() for s in exclude_ploshadkas]
            ploshadka_lower = self.df2[SYNTH_PLOSHADKA].astype(str).str.strip().str.lower()
            mask &= ~ploshadka_lower.isin(excl)
        df = self.df2[mask]
        rows = []
        for _, row in df.iterrows():
            qty = row[COLS_R2["qty"]]
            try:
                qty = int(qty) if pd.notna(qty) else 0
            except (ValueError, TypeError):
                qty = 0
            rows.append({
                "supp":        row[COLS_R2["supplier"]],
                "pid":         row[COLS_R2["pid"]],
                "prod_name":   row[COLS_R2["prod_name"]],
                "qty":         qty,
                "base_amount": row[COLS_R2["base_amount"]],
                "base_ccy":    row.get(COLS_R2["base_ccy"]),
                "prod_amount": row.get(COLS_R2["prod_amount"]),
                "prod_ccy":    row.get(COLS_R2["prod_ccy"]),
                "fx_rate":     row.get(COLS_R2["fx_rate"]),
            })
        return rows

    def _build_genba_lookup(self, ploshadka_filter) -> dict:
        if self.genba.empty:
            return {}
        filters = self._normalize_filter(ploshadka_filter)
        if not filters:
            return {}
        gp = self.genba[
            self.genba[COLS_GENBA["ploshadka"]].astype(str).str.strip().str.lower()
            .isin(filters)
        ]
        if gp.empty:
            return {}
        agg = gp.groupby(COLS_GENBA["pid"]).agg(
            qty=(COLS_GENBA["qty"], "sum"),
            cost=(COLS_GENBA["grand_total"], "sum"),
        )
        agg["price"] = agg["cost"] / agg["qty"]
        return agg["price"].to_dict()

    def _fx_lookup(self) -> dict:
        """Средний курс из R2 по валюте продукта (валюта → USD-множитель).

        Используется как фолбэк для расчёта цены в USD, когда у позиции
        пуст 'Курс фиксации' (так бывает в зоне 'Продажи б2б').

        Стартует со значений FX_FALLBACK_DEFAULT и переопределяет их средними
        курсами из тех R2-строк, где курс заполнен.
        """
        fx = dict(FX_FALLBACK_DEFAULT)
        if self.df2.empty:
            return fx
        rate_col = COLS_R2.get("fx_rate")
        ccy_col  = COLS_R2.get("prod_ccy")
        if rate_col not in self.df2.columns or ccy_col not in self.df2.columns:
            return fx
        sub = self.df2[[ccy_col, rate_col]].dropna()
        if sub.empty:
            return fx
        # средний курс по валюте (валюта → USD)
        means = sub.groupby(ccy_col)[rate_col].mean().to_dict()
        for ccy, rate in means.items():
            if rate and not pd.isna(rate):
                fx[str(ccy)] = float(rate)
        return fx

    @staticmethod
    def _compute_price(row, genba_lookup, fx_lookup=None):
        """Возвращает (unit_price, currency).

        Логика:
        1. CNY-поставщики: RUB-сумма из биллинга ÷ qty ÷ 11 → CNY.
        2. Genba: цена из genbaFile (USD), фолбэк — USD-сумма из биллинга.
        3. Прочие — берём итоговую валюту (SUPPLIER_CURRENCY или USD по умолчанию):
           a. Если итог = валюте продукта (prod_ccy) → цена = sum_prod / qty.
           b. Если итог = валюте базы (base_ccy через sum_base) → цена = sum_base / qty.
           c. Иначе нужна конвертация: sum_prod (в prod_ccy) * fx_rate → итог.
              fx_rate берётся из fx_lookup (средний курс из R2). Для USD = 1.
        """
        fx_lookup = fx_lookup or {}
        sg = row["supplier_group"]
        target_ccy = SUPPLIER_CURRENCY.get(sg, DEFAULT_CURRENCY)
        qty = row["qty"]

        # 1. CNY-поставщики
        if sg in CNY_SUPPLIERS:
            if qty > 0 and pd.notna(row.get("sum_rub")) and row["sum_rub"] != 0:
                return (row["sum_rub"] / qty / RUB_CNY_RATE, "CNY")
            return (None, "CNY")

        # 2. Genba: сначала из genbaFile, фолбэк на USD-сумму из биллинга
        if sg == "Genba":
            gpr = genba_lookup.get(row["pid"])
            if gpr is not None and pd.notna(gpr):
                return (gpr, "USD")

        if qty <= 0:
            return (None, target_ccy)

        prod_ccy = row.get("prod_ccy")
        sum_prod = row.get("sum_prod")
        sum_base = row.get("sum_base")

        # 3a. Итог в валюте продукта — берём prod-сумму без конвертации
        if pd.notna(sum_prod) and sum_prod != 0 and prod_ccy == target_ccy:
            return (sum_prod / qty, target_ccy)

        # 3b. Итог = USD и есть валидный base (он почти всегда в USD) → используем его
        if pd.notna(sum_base) and sum_base != 0 and target_ccy == "USD":
            return (sum_base / qty, target_ccy)

        # 3c. Конвертация: prod_amount → target через fx_lookup
        if pd.notna(sum_prod) and sum_prod != 0 and prod_ccy and target_ccy == "USD":
            rate = fx_lookup.get(str(prod_ccy))
            if rate:
                return (sum_prod / qty * rate, target_ccy)

        # 3d. Последний фолбэк — sum_base если он есть (старое поведение)
        if pd.notna(sum_base):
            return (sum_base / qty, target_ccy)

        return (None, target_ccy)

    # -----------------------------------------------------------------------
    # Сохранение в Excel в формате ЗАКУП (свод) — современный эстетичный вид
    # -----------------------------------------------------------------------
    def save_to_excel(
        self,
        agg: pd.DataFrame,
        ploshadka_key: str,
        out_path: str | Path,
        active_only: bool = True,
    ) -> Path:
        """Записывает агрегат в формат ЗАКУП (свод) с современным дизайном.

        Стилистика:
        — фирменный индиго в шапке таблицы (#4B4BFF, белый текст)
        — секции-поставщики с тёплым акцентом (#EEF0FF)
        — чередование строк продуктов (#FAFAFB)
        — приглушённые серые границы #E4E4EA
        — типографика Calibri 11pt
        — формулы Excel сохраняются (=C*D, =SUM(...)), чтобы можно было править на месте
        """
        if agg.empty:
            raise ValueError(f"Нет данных по площадке {ploshadka_key}")

        wb = Workbook()
        ws = wb.active
        ws.title = "ЗАКУП (свод)"

        # ---- Цвета и стили (фирменная палитра приложения) -----------------
        COLOR_HEADER_BG = "4B4BFF"        # индиго
        COLOR_HEADER_TEXT = "FFFFFF"
        COLOR_SECTION_BG = "EEF0FF"       # светло-индиго для строк-поставщиков
        COLOR_SECTION_TEXT = "26215C"     # тёмно-индиго
        COLOR_TOTAL_BG = "0A0A1F"         # тёмная итоговая строка
        COLOR_TOTAL_TEXT = "FFFFFF"
        COLOR_ALT_ROW = "FAFAFB"          # чередование
        COLOR_BORDER = "E4E4EA"           # приглушённый серый
        COLOR_TEXT = "0A0A1F"
        COLOR_TEXT_MUTED = "6B6B80"

        FONT_NAME = "Calibri"

        thin_side = Side(style="thin", color=COLOR_BORDER)
        thin_border = Border(left=thin_side, right=thin_side,
                             top=thin_side, bottom=thin_side)

        # ---- Шапка таблицы ------------------------------------------------
        headers = [
            "ID продукта",
            "Название",
            "Количество",
            "Цена закупа",
            "Валюта",
            "Себестоимость",
        ]

        # Лейбл площадки наверху
        ws.cell(row=1, column=1, value=f"Закуп · {ploshadka_key}")
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(name=FONT_NAME, bold=True, size=16, color=COLOR_TEXT)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        # Подзаголовок
        ws.cell(row=2, column=1, value="Свод по поставщикам в формате QuickBooks")
        ws.cell(row=2, column=1).font = Font(name=FONT_NAME, size=10, color=COLOR_TEXT_MUTED, italic=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

        # Шапка с колонками — строка 4
        HEADER_ROW = 4
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=HEADER_ROW, column=col, value=h)
            c.font = Font(name=FONT_NAME, bold=True, size=11, color=COLOR_HEADER_TEXT)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
            c.border = thin_border
        ws.row_dimensions[HEADER_ROW].height = 28

        # ---- Готовим данные ----------------------------------------------
        if active_only:
            agg = agg[agg["qty"] > 0]

        supplier_order = (
            agg.groupby("supplier_group")["cost"]
            .sum().sort_values(ascending=False).index
        )

        row_idx = HEADER_ROW + 1
        total_refs = []
        section_fill = PatternFill("solid", start_color=COLOR_SECTION_BG)
        alt_fill = PatternFill("solid", start_color=COLOR_ALT_ROW)

        for supp_name in supplier_order:
            if pd.isna(supp_name):
                continue
            sub = agg[agg["supplier_group"] == supp_name].sort_values("pid")
            if sub.empty:
                continue

            first_data = row_idx + 1
            last_data = row_idx + len(sub)

            # ---- Строка-заголовок секции -----------------------------
            for col in range(1, 7):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = section_fill
                cell.border = thin_border
                cell.font = Font(name=FONT_NAME, bold=True, size=11, color=COLOR_SECTION_TEXT)

            ws.cell(row=row_idx, column=2, value=supp_name)
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.cell(row=row_idx, column=3, value=f"=SUM(C{first_data}:C{last_data})")
            ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=row_idx, column=6, value=f"=SUM(F{first_data}:F{last_data})")
            ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[row_idx].height = 24
            total_refs.append(row_idx)
            row_idx += 1

            # ---- Строки продуктов с zebra-чередованием ---------------
            for i, (_, prod) in enumerate(sub.iterrows()):
                is_alt = i % 2 == 1
                # ID
                ws.cell(row=row_idx, column=1, value=int(prod["pid"]))
                ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
                # Название
                ws.cell(row=row_idx, column=2,
                        value=str(prod["prod_name"]) if pd.notna(prod["prod_name"]) else "")
                ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
                # Кол-во
                ws.cell(row=row_idx, column=3, value=int(prod["qty"]))
                ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right", vertical="center")
                # Цена
                if pd.notna(prod["unit_price"]):
                    ws.cell(row=row_idx, column=4, value=float(prod["unit_price"]))
                ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="right", vertical="center")
                # Валюта
                ws.cell(row=row_idx, column=5, value=str(prod["currency"]))
                ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center", vertical="center")
                # Себестоимость
                ws.cell(row=row_idx, column=6, value=f"=C{row_idx}*D{row_idx}")
                ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right", vertical="center")

                # Общие свойства строки
                for col in range(1, 7):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = Font(name=FONT_NAME, size=10, color=COLOR_TEXT)
                    cell.border = thin_border
                    if is_alt:
                        cell.fill = alt_fill
                ws.row_dimensions[row_idx].height = 20
                row_idx += 1

        # ---- TOTAL -------------------------------------------------------
        if total_refs:
            total_fill = PatternFill("solid", start_color=COLOR_TOTAL_BG)
            for col in range(1, 7):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = total_fill
                cell.font = Font(name=FONT_NAME, bold=True, size=12, color=COLOR_TOTAL_TEXT)
                cell.border = thin_border

            ws.cell(row=row_idx, column=2, value="ИТОГО")
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.cell(row=row_idx, column=3, value="=" + "+".join(f"C{r}" for r in total_refs))
            ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=row_idx, column=6, value="=" + "+".join(f"F{r}" for r in total_refs))
            ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[row_idx].height = 28

        # ---- Числовые форматы --------------------------------------------
        for r in range(HEADER_ROW + 1, row_idx + 1):
            ws.cell(row=r, column=3).number_format = '#,##0;-#,##0;""'
            ws.cell(row=r, column=4).number_format = '#,##0.0000;-#,##0.0000;""'
            ws.cell(row=r, column=6).number_format = '#,##0.00;-#,##0.00;""'

        # ---- Размеры колонок ---------------------------------------------
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 48
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 18

        # ---- Закрепляем шапку и убираем линии сетки ----------------------
        ws.freeze_panes = f"A{HEADER_ROW + 1}"
        ws.sheet_view.showGridLines = False

        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path

    # -----------------------------------------------------------------------
    # Плоское представление и CSV-экспорт
    # -----------------------------------------------------------------------
    def to_dataframe(
        self,
        agg: pd.DataFrame,
        ploshadka_key: str,
        active_only: bool = True,
    ) -> pd.DataFrame:
        """Возвращает плоский DataFrame для CSV-экспорта.

        Структура: одна строка на (поставщик × продукт). Поставщик дублируется
        в каждой строке — это удобнее для импорта в QuickBooks или сводных таблиц.
        Себестоимость — посчитанное число (не формула).
        """
        if agg.empty:
            return pd.DataFrame()

        df = agg.copy()
        if active_only:
            df = df[df["qty"] > 0]
        df = df[df["supplier_group"].notna()]

        # Сортируем: сначала по убыванию суммы поставщика, внутри — по pid
        supplier_totals = df.groupby("supplier_group")["cost"].sum()
        df["_supp_order"] = df["supplier_group"].map(supplier_totals)
        df = df.sort_values(["_supp_order", "supplier_group", "pid"],
                            ascending=[False, True, True])

        out = pd.DataFrame({
            "Площадка":      ploshadka_key,
            "Поставщик":     df["supplier_group"],
            "ID продукта":   df["pid"].astype("Int64"),
            "Название":      df["prod_name"].fillna(""),
            "Количество":    df["qty"].astype("Int64"),
            "Цена закупа":   df["unit_price"].round(4),
            "Валюта":        df["currency"],
            "Себестоимость": df["cost"].round(2),
        })
        return out.reset_index(drop=True)

    def save_to_csv(
        self,
        agg: pd.DataFrame,
        ploshadka_key: str,
        out_path: str | Path,
        active_only: bool = True,
    ) -> Path:
        """Сохраняет плоский CSV (UTF-8 с BOM, разделитель — запятая).

        BOM нужен чтобы Excel в Windows корректно открывал кириллицу.
        """
        df = self.to_dataframe(agg, ploshadka_key, active_only=active_only)
        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(out_path, index=False, encoding="utf-8-sig")
        return out_path


    # -----------------------------------------------------------------------
    # Инсайты для дашборда
    # -----------------------------------------------------------------------
    def insights(self, agg: pd.DataFrame) -> dict:
        """Возвращает агрегаты для инфографики по одной площадке.

        Структура:
            total_qty, total_cost, avg_price
            n_suppliers, n_products
            top_suppliers: [{name, qty, cost, n_products, pct_cost}]
            ccy_split: [{currency, qty, cost, pct_cost}]
            top_product: {pid, name, qty, cost}
            concentration: % крупнейшего поставщика
        """
        active = agg[agg["qty"] > 0]
        if active.empty:
            return {}

        total_qty = int(active["qty"].sum())
        total_cost = float(active["cost"].sum())

        # Топ поставщиков
        supp = (
            active.groupby("supplier_group")
            .agg(qty=("qty", "sum"), cost=("cost", "sum"), n_products=("pid", "count"))
            .sort_values("cost", ascending=False)
        )
        top_suppliers = [
            {
                "name": str(idx),
                "qty": int(row["qty"]),
                "cost": round(float(row["cost"]), 2),
                "n_products": int(row["n_products"]),
                "pct_cost": round(float(row["cost"]) / total_cost * 100, 1) if total_cost else 0,
            }
            for idx, row in supp.head(10).iterrows()
        ]

        # Валюты
        ccy = (
            active.groupby("currency")
            .agg(qty=("qty", "sum"), cost=("cost", "sum"))
            .sort_values("cost", ascending=False)
        )
        ccy_split = [
            {
                "currency": str(idx),
                "qty": int(row["qty"]),
                "cost": round(float(row["cost"]), 2),
                "pct_cost": round(float(row["cost"]) / total_cost * 100, 1) if total_cost else 0,
            }
            for idx, row in ccy.iterrows()
        ]

        # Топ-продукт
        top_prod_row = active.nlargest(1, "cost").iloc[0]
        top_product = {
            "pid": int(top_prod_row["pid"]),
            "name": str(top_prod_row["prod_name"]) if pd.notna(top_prod_row["prod_name"]) else "—",
            "qty": int(top_prod_row["qty"]),
            "cost": round(float(top_prod_row["cost"]), 2),
        }

        return {
            "total_qty": total_qty,
            "total_cost": round(total_cost, 2),
            "avg_price": round(total_cost / total_qty, 2) if total_qty else 0,
            "n_suppliers": int(active["supplier_group"].nunique()),
            "n_products": int(len(active)),
            "top_suppliers": top_suppliers,
            "ccy_split": ccy_split,
            "top_product": top_product,
            "concentration_pct": top_suppliers[0]["pct_cost"] if top_suppliers else 0,
            "concentration_supplier": top_suppliers[0]["name"] if top_suppliers else "",
        }

    def run_all(self, output_dir: str | Path) -> dict:
        """Прогоняет все доступные площадки, возвращает {ploshadka: путь_к_файлу}."""
        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        results = {}
        for key in PLOSHADKA_MAP:
            try:
                agg = self.aggregate(key)
                if agg.empty:
                    continue
                path = self.save_to_excel(
                    agg, key, out_dir / f"{key}_zakup_svod.xlsx"
                )
                results[key] = {
                    "path": str(path),
                    "qty": int(agg[agg["qty"] > 0]["qty"].sum()),
                    "cost": float(agg[agg["qty"] > 0]["cost"].sum()),
                    "suppliers": int(agg[agg["qty"] > 0]["supplier_group"].nunique()),
                }
            except Exception as e:
                results[key] = {"error": str(e)}
        return results
